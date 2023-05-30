"""
early_science_planning.py
Written by: Maren Cosens
Date: 3/31/23

Description: define constraints and optimize scheduling of early science target observations

Python version: 3.9

Notes: currently set up to use MoonSeparationConstraint and MoonIlluminationConstraint; can change to custom MoonSkyBrightnessConstraint by switching what's commented out

To-do: add stellar pops early science targets

Run as:
[] ipython
[] %run early_science_planning start_date(YYY-MM-DD) end_date(YYY-MM-DD) plot?True/False
"""
##import packages
from astropy.coordinates import SkyCoord, get_moon
from astroplan import FixedTarget, Observer, moon_phase_angle
from astropy.time import Time, TimezoneInfo
from astroplan import Constraint, AirmassConstraint, AtNightConstraint, MoonSeparationConstraint, MoonIlluminationConstraint
from astroplan import observability_table, ObservingBlock
from astroplan.plots import plot_airmass, plot_sky
from astroplan.utils import time_grid_from_range
from astroplan.scheduling import SequentialScheduler, PriorityScheduler, Schedule
import matplotlib.pyplot as plt
from astropy.table import Table
import astropy.units as u
import numpy as np
import os, copy, sys
from openpyxl import Workbook
from openpyxl.styles import Alignment

##read inputs from command line
#convert input 'True'/'False' to boolean; from https://stackoverflow.com/a/21732183
def str_to_bool(s):
    if s == 'True':
        return True
    elif s == 'False':
        return False
    else:
        raise ValueError("Cannot covert {} to a bool".format(s))
plotting = str_to_bool(sys.argv[-1])
end_date = str(sys.argv[-2])+" 06:00"
start_date = str(sys.argv[-3])+" 06:00"

#define observatory
lvmi = Observer.at_site('lco') #set location of observatory (astropy has coordinates/elevation for LCO already)
utc_minus_four_hours = TimezoneInfo(utc_offset=-4*u.hour)
time_range = Time([start_date, end_date])
date_array=np.arange(Time(start_date), Time(end_date), 24*u.hour)

##define targets; currently test with one, add capability to read/write from/to table "mini_proposal_targets_brightness.txt"
target_table = Table.read('mini_proposal_targets_brightness.txt', format='ascii')
targets = [FixedTarget(coord=SkyCoord(ra=RA, dec=Dec, frame='fk5', unit=(u.hourangle, u.deg)), name=target) for target, RA, Dec in target_table["target", "RA", "Dec"]]

##place constraints on targets
#general constraints for all targets
constraints = [AirmassConstraint(2, ), AtNightConstraint.twilight_civil(), MoonSeparationConstraint(min=45*u.deg)] #may not need MoonSeparationConstraint, but probably good for all targets to not be too close to the moon

##assign moon illumination constraint with bright/dark/grey commands based on quantiles of ha brightness (option 2)
#bright illumination > 65%
#grey 25% < illumination < 65%
#dark illumination < 25%
max_target_brightness = np.nanmax(target_table['mean_ha_brightness'])
scaled_target_brightness = target_table['mean_ha_brightness']/max_target_brightness
constraints_indiv_targets=[]
brightness_order = target_table.argsort("mean_ha_brightness") #returns sorted indices, faintest > brightest

for i in range(len(targets)):
    constraints_list=copy.copy(constraints)
    if i in brightness_order[:round(len(targets)/4)]:
        constraints_list.append(MoonIlluminationConstraint(max=.25)) #dark time
    elif i in brightness_order[round(len(targets)/4):round(len(targets)*(13/20))]:
        constraints_list.append(MoonIlluminationConstraint(max=.65)) #dark or grey time
    else:
        constraints_list.append(MoonIlluminationConstraint(max=1)) #any time
    constraints_indiv_targets.append(constraints_list)


##custom constraint for sky brightness combining MoonIlluminationConstraint and MoonSeparationConstraint
class MoonSkyBrightnessConstraint(Constraint):
    """
    Constraint on the additional sky brightness from the moon calculated via serparation and illumination
    """
    def __init__(self, min=None, max=None, boolean_constraint=True):
        """
        min : `~astropy.units.Quantity` or `None` (optional)
            Minimum acceptable sky brightness added by the moon. `None`
            indicates no limit.
        max : `~astropy.units.Quantity` or `None` (optional)
            Maximum acceptable sky brightness added by the moon. `None`
            indicates no limit.
        """
        self.min = min if min is not None else 100
        self.max = max if max is not None else -5
        self.boolean_constraint = boolean_constraint

    def compute_constraint(self, times, observer, targets):
            a=moon_phase_angle(times) #[radians]
            z_moon = 90*u.deg - observer.moon_altaz(times).alt #moon zenith angle [deg]
            z_target = 90*u.deg - observer.altaz(times, targets).alt #targets zenith angle [deg]
            moon = get_moon(times, location=observer.location)
            rho =  moon.separation(targets) #target/moon seperation [degrees]
            k = 0.172 #extinction coefficient [mag/airmass]; using rough estimate
            V_sky = 21.4 #average dark sky brightness [mag]; rough estimate, refine for LCO
            ##Equations from Krisciunas & Schaefer 1991
            I = 10**(-0.4*(3.84+0.026*np.abs(a.value)+4e-9*a.value**4)) #Eq 20
            f_rho = 10**5.36 * (1.06*np.cos(rho.value)**2) + 10**(6.15-rho.value/40) #Eq 21
            X_target = (1-0.96*np.sin(z_target)**2)**(-0.5) #Eq 3
            X_moon = (1-0.96*np.sin(z_moon)**2)**(-0.5) #Eq 3
            B_moon = f_rho * I * 10**(-0.4*k*X_moon) *(1-10**(-0.4*k*X_target)) #Eq 15
            B_0 = 34.08*np.exp(20.7233-0.92104*V_sky) #Eq 1; dark sky brightness [nanoLamberts]
            Delta_V = -2.5*np.log10((B_moon + B_0)/B_0) #mag/sec^2

            if self.boolean_constraint:
                mask = ((self.max < Delta_V) & (Delta_V < self.min))
                return mask

            # if we want to return a non-boolean score
            else:
                # rescale the vega_separation values so that they become
                # scores between zero and one
                rescale = min_best_rescale(Delta_V, self.max,
                                       self.min, less_than_min=0)
                return rescale

#use custom MoonSkyBrightnessConstraint instead of MoonSeparationConstraint and MoonIlluminationConstraint with allowed brightness based on target ha brightness
'''
for i in range(len(targets)):
    constraints_list=copy.copy(constraints)
    if i in brightness_order[:round(len(targets)/4)]:
        constraints_list.append(MoonSkyBrightnessConstraint(max=-2)) #sign due to magnitudes
    elif i in brightness_order[round(len(targets)/4):round(len(targets)*(13/20))]:
        constraints_list.append(MoonSkyBrightnessConstraint(max=-3.5)) #min=0 allows it to be observable during dark or grey time
    else:
        constraints_list.append(MoonSkyBrightnessConstraint(max=-5))
    constraints_indiv_targets.append(constraints_list)
'''

##generate quick table of whether they are observable in the time window and for what fraction
observability_table = observability_table(constraints, lvmi, targets, time_range=time_range)
print(observability_table)

def target_observability(target, constraints, observe_time, figure_path):
    """
    Evaluate/plot the observability of the specified target given the constraints
    Inputs:
    target : single astroplan target object
    constraints: list of Astroplan constraints to be evaluated
    observe_time : time_grid over which to evaluate the constraints
    figure_path : directory in which to save the figure
    Returns:
    observability_grid : grid of observability for each constraint
    observability_plot : plots value of each constraint for the target over the specified range of time
    """
    n_constraints=len(constraints)
    observability_grid = np.zeros((len(constraints), len(observe_time)))
    times_observable = np.zeros(len(observe_time))
    for i, constraint in enumerate(constraints):
        # Evaluate each constraint
        observability_grid[i, :] = constraint(lvmi, target, times=observe_time)
    for i in range(len(observe_time)):
        times_observable[i] = np.sum(observability_grid[:,i])/n_constraints
        if times_observable[i]<1:
            times_observable[i]=0

    # Create plot showing observability of the target:
    extent = [-0.5, -0.5+len(observe_time), -0.5, n_constraints-0.5]
    fig=plt.figure(target.name, figsize=(14,5))
    ax = fig.subplots()
    ax.set_title(target.name)
    ax.imshow(observability_grid, extent=extent)
    ax.set_yticks(range(0, n_constraints))
    ylabels=[c.__class__.__name__ for c in constraints]
    ax.set_yticklabels(ylabels[::-1])
    ax.set_xticks(range(len(observe_time)))
    ax.set_xticklabels([t.to_datetime(timezone=utc_minus_four_hours).strftime("%H:%M") for t in observe_time])
    ax.set_xticks(np.arange(extent[0], extent[1]), minor=True)
    ax.set_yticks(np.arange(extent[2], extent[3]), minor=True)
    ax.grid(which='minor', color='w', linestyle='-', linewidth=2)
    ax.tick_params(axis='x', which='minor', bottom='off')
    plt.setp(ax.get_xticklabels(), rotation=30, ha='right')
    ax.tick_params(axis='y', which='minor', left='off')
    ax.set_xlabel('Time on {0} CLT'.format(observe_time[0].to_datetime(timezone=utc_minus_four_hours).date()))
    fig.subplots_adjust(left=0.148, right=0.979, top=0.99, bottom=0.1)
    plt.savefig(str(figure_path)+'/'+str(target.name)+'_constraint_plot.png', bbox_inches='tight')
    plt.close()

    return observability_grid, times_observable

##check observability for each date
#openpyxl to save observability list to one file
wb = Workbook()
ws1=wb.active #get active sheet (first one)
ws1.title='README'
ws1["A1"].value = "List of observable early science targets organized by date and time"
ws1["A2"].value = "Sheets correspond to individual nights with the date being the start of the night (CST)"
ws1["A3"].value = "Target observability is recorded in 1 hour blocks with the local time (CST) in row 1"
ws1["A4"].value = "A list of all observable targets for a given time is found in row 2."
#look through hour blocks for each night
time_resolution = 1 * u.hour
for time in date_array:
    ##define obs date
    sunset_tonight = lvmi.sun_set_time(time, which='nearest')
    sunrise_tonight = lvmi.sun_rise_time(time, which='nearest')
    obs_date=sunset_tonight.to_datetime(timezone=utc_minus_four_hours).strftime("%b-%d-%Y")
    figure_path = "obs_constraint_plots/"+str(obs_date)
    # Check whether the specified path exists or not
    isExist = os.path.exists(figure_path)
    if not isExist:
        # Create a new directory because it does not exist
        os.makedirs(figure_path)

    ##generate plots of observability
    delta_t = sunrise_tonight - sunset_tonight
    observe_time = time_grid_from_range([sunset_tonight, sunrise_tonight], time_resolution=time_resolution)#sunset_tonight + delta_t*np.linspace(0, 1, 75)

    obs_grids = [] # array of booleans for each time window/constraint/target
    obs_times = [] #array of booleans for each time window/target specifying whether all constraints are met
    for i in range(len(targets)):
        #print(targets[i].name)
        grids, windows = target_observability(targets[i], constraints_indiv_targets[i], observe_time, figure_path)
        obs_grids.append(grids)
        obs_times.append(windows)

    ##generate dictionary of targets observable in each time window
    time_list=[t.strftime("%H:%M") for t in observe_time.to_datetime(timezone=utc_minus_four_hours)]
    total_observable_targets = dict()
    for i in range(len(time_list)):
        observable_targets=[]
        for j in range(len(targets)):
            if obs_times[j][i]==1:
                observable_targets.append(targets[j].name)
                total_observable_targets[time_list[i]]=observable_targets
                #print(total_observable_targets)
    #save to one file with a sheet for each night
    ws=wb.create_sheet(str(obs_date)) #add sheet at end
    ws.title = str(obs_date) #update sheet title to start of night date
    #loop through all times in list to add to sheet
    for j in range(0,len(time_list)):
        ws.cell(row=1, column=j+1).value = time_list[j]
        if time_list[j] in total_observable_targets:
            ws.cell(row=2, column=j+1).value = '\n'.join(total_observable_targets[time_list[j]]) #save list of targets observable as single string in cell
            ws.cell(row=2, column=j+1).alignment = Alignment(wrapText=True)

wb.save("EarlyScienceObservability.xlsx")
